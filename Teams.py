# code creates Qtablewidget for teams and updates owners with purchased player

import csv
import openpyxl

from PyQt5.QtGui import QFont, QColor
from PyQt5.QtWidgets import QTableWidget, QWidget, QVBoxLayout, QTableWidgetItem, QCheckBox, QFileDialog, QMessageBox

HEADERS = ['Remaining Money', 'Quarterback', 'Running Back', 'Running Back 2', 'Wide Receiver', 'Wide Receiver 2',
           'Tight End', 'Flex', 'Kicker', 'Defense', 'Bench 1', 'Bench 2', 'Bench 3', 'Bench 4', 'Bench 5']

HEADER_DICT = {"QB": 1, "RB": 2, "WR": 4, "TE": 6, "K": 8, "DEF": 9}


# initial set up of QTableWidget, accepts choice, num and arguments from MainWindow.py
class TeamView(QWidget):
    def __init__(self, choice, num, *args, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Team View')
        self.setGeometry(100, 100, 1350, 650)

        self.updateComp = True

        self.rowCount = num
        self.teamNames = []
        for name in args[0]:
            self.teamNames.append(name)

        self.teamTbl = QTableWidget()
        self.teamTbl.setRowCount(self.rowCount * 2)
        self.teamTbl.setColumnCount(len(HEADERS))

        # Column set up
        self.teamTbl.setHorizontalHeaderLabels(HEADERS)

        for i in range(len(HEADERS)):
            self.teamTbl.setColumnWidth(i, 180)

        # row set up
        r = 0
        for name in self.teamNames:
            self.teamTbl.setVerticalHeaderItem(r, QTableWidgetItem(name))
            r += 1
            self.teamTbl.setVerticalHeaderItem(r, QTableWidgetItem('$'))
            r += 1

        self.tblSetUp()
        # loads saved table
        if choice == 'Load':
            self.loadTable()

        self.teamTbl.setStyleSheet("""
            QHeaderView::section {
                background: rgb(192, 192, 192);
                color: rgb(0, 0, 0);
                border: 1px solid black;
                font: bold 12pt;
                text-align: right;
            }
            QTableWidget{
                color: rgb(0, 0, 0);
            }
            QTableWidget QTableCornerButton::section {
                background: rgb(192, 192, 192)
            }
        """)

        self.teamTbl.itemChanged.connect(self.itemChanged)

        self.vbox = QVBoxLayout()
        self.vbox.addWidget(self.teamTbl)
        self.setLayout(self.vbox)

        self.setVisible(False)
        self.saveTeams()

    # definition sets up the horizontal and vertical headers of table
    def tblSetUp(self):
        totRows = self.rowCount * 2
        font = QFont()
        font.setPointSize(12)
        font.setBold(True)

        r = 0
        while r < totRows:
            for c in range(self.teamTbl.columnCount()):
                self.teamTbl.setItem(r, c, QTableWidgetItem(''))
                self.teamTbl.item(r, c).setBackground(QColor(240, 241, 242))
                self.teamTbl.item(r, c).setFont(font)
            self.teamTbl.setCellWidget(r, 0, QCheckBox("Paid"))
            # self.teamTbl.cellWidget(r, 0).setFont(font)
            self.teamTbl.cellWidget(r, 0).setStyleSheet("""
                QCheckBox {
                    background: rgb(240, 241, 242);
                    color: rgb(0, 0, 128);
                    font: 12pt;
                }
            """)
            r += 1
            for c in range(self.teamTbl.columnCount()):
                self.teamTbl.setItem(r, c, QTableWidgetItem(''))
                self.teamTbl.item(r, c).setBackground(QColor(125, 140, 155))
                self.teamTbl.item(r, c).setFont(font)
            self.teamTbl.item(r, 0).setText('300')
            # self.teamTbl.item(r, 0).setForeground(QColor(50, 75, 0))
            r += 1

    # checks player position and owner, then runs code based on the position and places player in row based on owner
    def updateTable(self, owner, player, amt, position):
        self.updateComp = True
        row = 0
        for i in range(self.teamTbl.rowCount()):
            own = self.teamTbl.verticalHeaderItem(i)
            if own.text() == owner:
                row = i
            else:
                pass

        if position == 'QB':
            self.updateQB(player, amt, row)
        elif position == 'RB':
            self.updateRB(player, amt, row)
        elif position == 'WR':
            self.updateWR(player, amt, row)
        elif position == 'TE':
            self.updateTE(player, amt, row)
        elif position == 'K':
            self.updateKicker(player, amt, row)
        else:
            self.updateDEF(player, amt, row)

        self.updateMoney(row + 1)
        self.saveTable()

    # definitions place player in open slot, if none are left, message box pops up
    def updateQB(self, player, amt, r):
        if not self.teamTbl.item(r, 1).text():
            self.teamTbl.item(r, 1).setText(player)
            self.teamTbl.item(r + 1, 1).setText(amt)
        elif not self.teamTbl.item(r, 14).text():
            self.updateBN(player, amt, r)
        elif not self.teamTbl.item(r, 7).text():
            self.updateFlex(player, amt, r)
        elif not self.teamTbl.item(r, 8).text():
            self.updateKicker(player, amt, r)
        elif not self.teamTbl.item(r, 9).text():
            self.updateDEF(player, amt, r)
        else:
            self.spotsFilled()

    def updateRB(self, player, amt, r):
        if not self.teamTbl.item(r, 2).text():
            self.teamTbl.item(r, 2).setText(player)
            self.teamTbl.item(r + 1, 2).setText(amt)
        elif not self.teamTbl.item(r, 3).text():
            self.teamTbl.item(r, 3).setText(player)
            self.teamTbl.item(r + 1, 3).setText(amt)
        elif not self.teamTbl.item(r, 7).text():
            self.updateFlex(player, amt, r)
        elif not self.teamTbl.item(r, 14).text():
            self.updateBN(player, amt, r)
        elif not self.teamTbl.item(r, 8).text():
            self.updateKicker(player, amt, r)
        elif not self.teamTbl.item(r, 9).text():
            self.updateDEF(player, amt, r)
        else:
            self.spotsFilled()

    def updateWR(self, player, amt, r):
        if not self.teamTbl.item(r, 4).text():
            self.teamTbl.item(r, 4).setText(player)
            self.teamTbl.item(r + 1, 4).setText(amt)
        elif not self.teamTbl.item(r, 5).text():
            self.teamTbl.item(r, 5).setText(player)
            self.teamTbl.item(r + 1, 5).setText(amt)
        elif not self.teamTbl.item(r, 7).text():
            self.updateFlex(player, amt, r)
        elif not self.teamTbl.item(r, 14).text():
            self.updateBN(player, amt, r)
        elif not self.teamTbl.item(r, 8).text():
            self.updateKicker(player, amt, r)
        elif not self.teamTbl.item(r, 9).text():
            self.updateDEF(player, amt, r)
        else:
            self.spotsFilled()

    def updateTE(self, player, amt, r):
        if not self.teamTbl.item(r, 6).text():
            self.teamTbl.item(r, 6).setText(player)
            self.teamTbl.item(r + 1, 6).setText(amt)
        elif not self.teamTbl.item(r, 7).text():
            self.updateFlex(player, amt, r)
        elif not self.teamTbl.item(r, 14).text():
            self.updateBN(player, amt, r)
        elif not self.teamTbl.item(r, 8).text():
            self.updateKicker(player, amt, r)
        elif not self.teamTbl.item(r, 9).text():
            self.updateDEF(player, amt, r)
        else:
            self.spotsFilled()

    def updateFlex(self, player, amt, r):
        if not self.teamTbl.item(r, 7).text():
            self.teamTbl.item(r, 7).setText(player)
            self.teamTbl.item(r + 1, 7).setText(amt)

    def updateKicker(self, player, amt, r):
        if not self.teamTbl.item(r, 8).text():
            self.teamTbl.item(r, 8).setText(player)
            self.teamTbl.item(r + 1, 8).setText(amt)
        elif not self.teamTbl.item(r, 14).text():
            self.updateBN(player, amt, r)
        elif not self.teamTbl.item(r, 9).text():
            self.updateDEF(player, amt, r)
        else:
            self.spotsFilled()

    def updateDEF(self, player, amt, r):
        if not self.teamTbl.item(r, 9).text():
            self.teamTbl.item(r, 9).setText(player)
            self.teamTbl.item(r + 1, 9).setText(amt)
        elif not self.teamTbl.item(r, 14).text():
            self.updateBN(player, amt, r)
        elif not self.teamTbl.item(r, 8).text():
            self.updateKicker(player, amt, r)
        else:
            self.spotsFilled()

    def updateBN(self, player, amt, r):
        if not self.teamTbl.item(r, 10).text():
            self.teamTbl.item(r, 10).setText(player)
            self.teamTbl.item(r + 1, 10).setText(amt)
        elif not self.teamTbl.item(r, 11).text():
            self.teamTbl.item(r, 11).setText(player)
            self.teamTbl.item(r + 1, 11).setText(amt)
        elif not self.teamTbl.item(r, 12).text():
            self.teamTbl.item(r, 12).setText(player)
            self.teamTbl.item(r + 1, 12).setText(amt)
        elif not self.teamTbl.item(r, 13).text():
            self.teamTbl.item(r, 13).setText(player)
            self.teamTbl.item(r + 1, 13).setText(amt)
        elif not self.teamTbl.item(0, 14).text():
            self.teamTbl.item(r, 14).setText(player)
            self.teamTbl.item(r + 1, 14).setText(amt)
        else:
            return None

    # displays message box if no spots are available. changes updatecomp to False
    def spotsFilled(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Warning)
        msg.setText("No available spots left. Returning player back to data set. "
                    "Please check Team View and confirm entries")
        msg.setWindowTitle("All Spots Filled")
        msg.setStandardButtons(QMessageBox.Close)
        self.updateComp = False
        msg.exec_()

    # definition accepts row number and then updates the money column
    def updateMoney(self, r):
        amt = 300

        c = 1
        totSpent = []
        while c in range(self.teamTbl.columnCount()):
            if self.teamTbl.item(r, c).text() == '':
                pass
            else:
                amtSpent = int(self.teamTbl.item(r, c).text())
                totSpent.append(amtSpent)
            c += 1

        newAmt = amt - sum(totSpent)
        self.teamTbl.item(r, 0).setText(str(newAmt))

    def getMoney(self, owner):
        amtLeft = 0
        for i in range(self.teamTbl.rowCount()):
            own = self.teamTbl.verticalHeaderItem(i)
            if own.text() == owner:
                amtLeft = int(self.teamTbl.item((i+1), 0).text())
            else:
                pass
        return amtLeft

    # if cell value is changed, function runs and gets the row # and updates money column
    def itemChanged(self, item):
        row = item.row()
        if row % 2 == 0:
            pass
        else:
            try:
                self.updateMoney(row)
            except ValueError:
                msg = QMessageBox()
                msg.setWindowTitle('Invalid Entry')
                msg.setIcon(QMessageBox.Warning)
                msg.setText(f"Numbers only. Please remove '{item.text()}' from cell.")
                msg.setStandardButtons(QMessageBox.Close)
                msg.exec_()

    # definition saves progress of the table to a csv file that can loaded when program starts
    def saveTable(self):
        with open('Saved/Table.csv', 'w', newline='') as outfile:
            writer = csv.writer(outfile)
            for row in range(self.teamTbl.rowCount()):
                rowData = []
                for col in range(self.teamTbl.columnCount()):
                    item = self.teamTbl.item(row, col)
                    if item is not None:
                        rowData.append(item.text())
                    else:
                        rowData.append('')
                writer.writerow(rowData)

    # definition saves team names to csv file that can be loaded when programs starts
    def saveTeams(self):
        with open('Saved/Team Names.csv', 'w', newline='') as outfile:
            wr = csv.writer(outfile, quoting=csv.QUOTE_ALL)
            wr.writerow(self.teamNames)

    # definition loads a saved table from csv file, user chooses to load at start of program
    def loadTable(self):
        with open('Saved/Table.csv', 'r') as file:
            reader = csv.reader(file)
            data = []
            for row in reader:
                data.append(row)

            for r in range(self.teamTbl.rowCount()):
                k = 0
                for c in range(self.teamTbl.columnCount()):
                    self.teamTbl.item(r, c).setText(data[r][k])
                    k += 1

    # exports table to spreadsheet as .xlsx file
    def export2excel(self):
        file = QFileDialog.getSaveFileName(self, 'Save File', 'c:\\', "Excel Files (*.xlsx)")
        if file[0] != '':
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = 'Results'

            # iterates table headers and assigns to cells in first row
            c = 2
            for col in HEADERS:
                cell = sheet.cell(row=1, column=c)
                cell.value = col
                c += 1

            # iterates team names and assigns to cells in first column
            r = 2
            for name in self.teamNames:
                cell = sheet.cell(row=r, column=1)
                cell.value = name
                r += 1
                moneyCell = sheet.cell(row=r, column=1)
                moneyCell.value = '$'
                r += 1

            # iterates through Qtablewidget and assigns values to cells in spreadsheet
            r = 1
            for row in range(self.teamTbl.rowCount()):
                r += 1
                c = 2
                for col in range(self.teamTbl.columnCount()):
                    item = self.teamTbl.item(row, col)
                    cell = sheet.cell(row=r, column=c)
                    if item is not None:
                        cell.value = item.text()
                    else:
                        cell.value = ''
                    c += 1

            wb.save(file[0])
        else:
            return

# if __name__ == '__main__':
#     app = QApplication(sys.argv)
#     win = TeamView()
#     win.show()
#     sys.exit(app.exec_())
